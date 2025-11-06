import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from tkcalendar import DateEntry
import mysql.connector
from mysql.connector import Error
import re
from PIL import Image, ImageTk
import os
from datetime import datetime, timedelta
from io import BytesIO
# Librerías para exportación
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.units import inch


# ----------------------------------------
# CLASE PRINCIPAL: DiarioEmocionesApp
# ----------------------------------------
class DiarioEmocionesApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Diario de Emociones 🧠❤️")
        self.root.geometry("1000x800")
        self.root.configure(bg="#faf3e0")

        self.configurar_favicon()

        # Conexión a MySQL
        self.conn = self.crear_conexion_mysql()

        # Configurar estilo
        self.configurar_estilo()

        # Crear interfaz modular
        self.crear_notebook()
        self.crear_pestañas()
        self.crear_footer()

    def configurar_favicon(self):
        try:
            self.root.iconbitmap("favicon.ico")
            print("✅ Favicon ICO cargado correctamente")
        except Exception as e:
            print(f"❌ Error cargando favicon ICO: {e}")

    def crear_conexion_mysql(self):
        try:
            conexion = mysql.connector.connect(
                host='localhost',
                database='diario_emociones',
                user='root',
                password=''
            )
            if conexion.is_connected():
                print("✅ Conexión a base de datos exitosa")
                return conexion
        except Exception as e:
            print(f"❌ Error de conexión: {e}")
            return None

    def configurar_estilo(self):
        style = ttk.Style()
        style.theme_use('clam')
        style.configure("TNotebook", background="#faf3e0")
        style.configure("TNotebook.Tab",
                        background="#e0d3c1",
                        foreground="#5a4a42",
                        font=("Arial", 11, "bold"),
                        padding=[12, 6])
        style.map("TNotebook.Tab",
                  background=[("selected", "#d4a5a5")],
                  foreground=[("selected", "white")])

        style.configure("Treeview",
                        background="#ffffff",
                        foreground="#5a4a42",
                        rowheight=25,
                        fieldbackground="#ffffff")
        style.configure("Treeview.Heading",
                        background="#d4a5a5",
                        foreground="white",
                        font=("Arial", 10, "bold"))
        style.map("Treeview",
                  background=[('selected', '#e77f67')])

    def crear_notebook(self):
        self.notebook = ttk.Notebook(self.root)
        self.notebook.pack(expand=True, fill="both", padx=20, pady=20)

    def crear_pestañas(self):
        # Módulo 1: Usuarios
        self.tab_usuarios = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_usuarios, text=" 👤 Usuarios ")
        self.crear_formulario_usuarios()

        # Módulo 2: Emociones
        self.tab_emociones = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_emociones, text=" 😊 Emociones ")
        self.crear_formulario_emociones()

        # Módulo 3: Entradas
        self.tab_entradas = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_entradas, text=" 📖 Entradas ")
        self.crear_formulario_entradas()

        # Módulo 4: Reportes
        self.tab_reportes = ttk.Frame(self.notebook)
        self.notebook.add(self.tab_reportes, text=" 📈 Reportes ")
        self.crear_formulario_reportes()

    # -------- VALIDACIONES DE LOS CAMPOS ----------
    def validar_id_usuario(self):
        user_id = self.usuario_id_entry.get()
        if user_id and not user_id.isdigit():
            messagebox.showerror("Error", "ID Usuario debe ser un número")
            return False
        return True

    def validar_id_emocion(self):
        emocion = self.emocion_id_entry.get()
        if emocion and not emocion.isdigit():
            messagebox.showerror("Error", "ID Emoción debe ser un número")
            return False
        return True

    def validar_id_entrada(self):
        entrada_id = self.entrada_id_entry.get()
        if entrada_id and not entrada_id.isdigit():
            messagebox.showerror("Error", "ID Entrada debe ser un número")
            return False
        return True

    def validar_usuario_id_entrada(self):
        usuario_id = self.entrada_usuario_id_entry.get()
        if usuario_id and not usuario_id.isdigit():
            messagebox.showerror("Error", "ID Usuario en Entradas debe ser un número")
            return False
        return True

    # ---------- VALIDACIONES DE TEXTO --------------
    def validar_texto_username(self, texto):
        if not texto:
            messagebox.showerror("Error", "Username es obligatorio")
            return False
        if not (3 <= len(texto) <= 50):
            messagebox.showerror("Error", "Username debe tener entre 3 y 50 caracteres")
            return False
        return True

    def validar_texto_nombre_emocion(self, texto):
        if not texto:
            messagebox.showerror("Error", "Nombre de la emoción es obligatorio")
            return False
        if not (1 <= len(texto) <= 50):
            messagebox.showerror("Error", "Nombre debe tener entre 1 y 50 caracteres")
            return False
        return True

    def validar_texto_password(self, texto):
        if not texto:
            messagebox.showerror("Error", "Password es obligatoria")
            return False
        if len(texto) < 6:
            messagebox.showerror("Error", "Password debe tener al menos 6 caracteres")
            return False
        return True

    def validar_entrada(self, texto):
        if not texto or texto.strip() == "":
            messagebox.showerror("Error", "El texto de la entrada es obligatorio")
            return False
        return True

    # -------- VALIDACIONES DE EMAIL---------
    def validar_email_usuario(self, email):
        patron = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
        if not email:
            messagebox.showerror("Error", "Email es obligatorio")
            return False
        if not re.match(patron, email):
            messagebox.showerror("Error", "Formato de email incorrecto")
            return False
        return True

    # --------- VALIDACIONES Y PROCESAMIENTO DE IMAGEN ---
    def validar_imagen(self, filepath):
        if not filepath:
            return True
        try:
            img = Image.open(filepath)
            if img.format not in ["JPEG", "PNG", "GIF"]:
                messagebox.showerror("Error", "Formato de imagen no soportado. Use JPG, PNG o GIF")
                return False
            if img.width > 1920 or img.height > 1080:
                messagebox.showwarning("Advertencia", "Imagen muy grande. Se recomienda resolución menor.")
                return True
            return True
        except Exception as e:
            messagebox.showerror("Error", f"Imagen Inválida: {e}")
            return False

    def procesar_imagen(self, filepath, max_width=800, max_height=800):
        """Redimensiona y optimiza la imagen usando Pillow"""
        try:
            if not filepath:
                return None
            
            img = Image.open(filepath)
            
            # Redimensionar si es muy grande
            if img.width > max_width or img.height > max_height:
                img.thumbnail((max_width, max_height), Image.Resampling.LANCZOS)
                messagebox.showinfo("Imagen procesada", f"Imagen redimensionada a {img.width}x{img.height}px")
            
            # Convertir a RGB si es necesario (para compatibilidad)
            if img.mode in ('RGBA', 'LA', 'P'):
                background = Image.new('RGB', img.size, (255, 255, 255))
                if img.mode == 'P':
                    img = img.convert('RGBA')
                background.paste(img, mask=img.split()[-1] if 'A' in img.mode else None)
                img = background
            
            # Guardar en BytesIO para almacenar en DB o usar
            output = BytesIO()
            img.save(output, format='JPEG', quality=85, optimize=True)
            output.seek(0)
            
            return output.getvalue()
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al procesar imagen: {e}")
            return None

    def mostrar_preview_imagen(self, filepath, label_widget, max_size=(150, 150)):
        """Muestra preview de la imagen en un label"""
        try:
            if filepath and os.path.exists(filepath):
                img = Image.open(filepath)
                img.thumbnail(max_size, Image.Resampling.LANCZOS)
                photo = ImageTk.PhotoImage(img)
                label_widget.config(image=photo)
                label_widget.image = photo  # Mantener referencia
                return True
        except Exception as e:
            print(f"Error mostrando preview: {e}")
        return False

    # -------- FUNCION PARA SELECCIONAR IMAGEN -----------
    def seleccionar_imagen(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar imagen de perfil",
            filetypes=[("Image Files", "*.jpg;*.jpeg;*.png;*.gif")]
        )
        if filepath and self.validar_imagen(filepath):
            self.imagen_path.set(filepath)
            # Mostrar preview
            self.mostrar_preview_imagen(filepath, self.preview_label_usuario)

    def seleccionar_imagen_emocion(self):
        filepath = filedialog.askopenfilename(
            title="Seleccionar imagen para la Emoción",
            filetypes=[("Image Files", "*.jpg;*.jpeg;*.png;*.gif")]
        )
        if filepath and self.validar_imagen(filepath):
            self.imagen_emocion_path.set(filepath)
            # Mostrar preview
            self.mostrar_preview_imagen(filepath, self.preview_label_emocion)

    # -------- FUNCIONES PARA CARGAR DATOS EN TABLAS (USANDO STORED PROCEDURES) -----------
    def cargar_usuarios_tabla(self):
        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_listar_usuarios')
            
            for result in cursor.stored_results():
                usuarios = result.fetchall()

            for item in self.tabla_usuarios.get_children():
                self.tabla_usuarios.delete(item)

            for usuario in usuarios:
                self.tabla_usuarios.insert("", "end", values=usuario)
        except Error as e:
            messagebox.showerror("Error", f"Error al cargar usuarios: {e}")

    def cargar_emociones_tabla(self):
        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_listar_emociones')
            
            for result in cursor.stored_results():
                emociones = result.fetchall()

            for item in self.tabla_emociones.get_children():
                self.tabla_emociones.delete(item)

            for emocion in emociones:
                self.tabla_emociones.insert("", "end", values=emocion)
        except Error as e:
            messagebox.showerror("Error", f"Error al cargar emociones: {e}")

    def cargar_entradas_tabla(self):
        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_listar_entradas')
            
            for result in cursor.stored_results():
                entradas = result.fetchall()

            for item in self.tabla_entradas.get_children():
                self.tabla_entradas.delete(item)

            for entrada in entradas:
                self.tabla_entradas.insert("", "end", values=entrada)
        except Error as e:
            messagebox.showerror("Error", f"Error al cargar entradas: {e}")

    # -------- FUNCIONES PARA SELECCIONAR FILAS EN TABLAS -----------
    def seleccionar_usuario_tabla(self, event):
        selected = self.tabla_usuarios.focus()
        if selected:
            values = self.tabla_usuarios.item(selected, 'values')
            self.usuario_id_entry.delete(0, tk.END)
            self.username_entry.delete(0, tk.END)
            self.email_entry.delete(0, tk.END)
            self.password_entry.delete(0, tk.END)

            self.usuario_id_entry.insert(0, values[0])
            self.username_entry.insert(0, values[1])
            self.email_entry.insert(0, values[2])

    def seleccionar_emocion_tabla(self, event):
        selected = self.tabla_emociones.focus()
        if selected:
            values = self.tabla_emociones.item(selected, 'values')
            self.emocion_id_entry.delete(0, tk.END)
            self.nombre_emocion_entry.delete(0, tk.END)

            self.emocion_id_entry.insert(0, values[0])
            self.nombre_emocion_entry.insert(0, values[1])

    def seleccionar_entrada_tabla(self, event):
        selected = self.tabla_entradas.focus()
        if selected:
            values = self.tabla_entradas.item(selected, 'values')
            self.entrada_id_entry.delete(0, tk.END)
            self.entrada_id_entry.insert(0, values[0])

    # ============================================
    # FUNCIONES DE EXPORTACIÓN
    # ============================================
    
    def exportar_usuarios_excel(self):
        """Exporta usuarios a Excel con formato profesional"""
        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_listar_usuarios')
            
            datos = None
            for result in cursor.stored_results():
                datos = result.fetchall()
            
            if not datos:
                messagebox.showinfo("Sin datos", "No hay usuarios para exportar")
                return
            
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not archivo:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Usuarios"
            
            header_fill = PatternFill(start_color="D4A5A5", end_color="D4A5A5", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            headers = ["ID", "Username", "Email", "Fecha Registro"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, row_data in enumerate(datos, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 20
            
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Usuarios exportados correctamente a:\n{archivo}")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar usuarios: {e}")

    def exportar_emociones_excel(self):
        """Exporta emociones a Excel"""
        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_listar_emociones')
            
            datos = None
            for result in cursor.stored_results():
                datos = result.fetchall()
            
            if not datos:
                messagebox.showinfo("Sin datos", "No hay emociones para exportar")
                return
            
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"emociones_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not archivo:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Emociones"
            
            header_fill = PatternFill(start_color="88C9A1", end_color="88C9A1", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            headers = ["ID", "Nombre", "Emoji"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, row_data in enumerate(datos, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 25
            ws.column_dimensions['C'].width = 15
            
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Emociones exportadas correctamente")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar emociones: {e}")

    def exportar_entradas_excel(self, fecha_inicio=None, fecha_fin=None, usuario_id=None):
        """Exporta entradas a Excel con filtros opcionales"""
        try:
            cursor = self.conn.cursor()
            
            query = """
                SELECT e.id, u.username, e.fecha, e.texto 
                FROM entradas e 
                JOIN usuarios u ON e.usuario_id = u.id 
                WHERE 1=1
            """
            params = []
            
            if fecha_inicio:
                query += " AND e.fecha >= %s"
                params.append(fecha_inicio)
            if fecha_fin:
                query += " AND e.fecha <= %s"
                params.append(fecha_fin)
            if usuario_id:
                query += " AND e.usuario_id = %s"
                params.append(usuario_id)
            
            query += " ORDER BY e.fecha DESC"
            
            cursor.execute(query, params)
            datos = cursor.fetchall()
            
            if not datos:
                messagebox.showinfo("Sin datos", "No hay entradas para exportar con los filtros aplicados")
                return
            
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"entradas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not archivo:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Entradas"
            
            header_fill = PatternFill(start_color="A2B9BC", end_color="A2B9BC", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            headers = ["ID", "Usuario", "Fecha", "Texto"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, row_data in enumerate(datos, 2):
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    if col_idx == 4:
                        cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws.column_dimensions['A'].width = 10
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 60
            
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Entradas exportadas: {len(datos)} registros")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar entradas: {e}")

    def exportar_usuarios_pdf(self):
        """Exporta usuarios a PDF con formato profesional"""
        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_listar_usuarios')
            
            datos = None
            for result in cursor.stored_results():
                datos = result.fetchall()
            
            if not datos:
                messagebox.showinfo("Sin datos", "No hay usuarios para exportar")
                return
            
            archivo = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=f"usuarios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            
            if not archivo:
                return
            
            doc = SimpleDocTemplate(archivo, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#B86D6D'),
                spaceAfter=30,
                alignment=1
            )
            title = Paragraph("📋 Reporte de Usuarios", title_style)
            elements.append(title)
            
            fecha_style = ParagraphStyle(
                'Fecha',
                parent=styles['Normal'],
                fontSize=10,
                textColor=colors.grey,
                alignment=1
            )
            fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}", fecha_style)
            elements.append(fecha)
            elements.append(Spacer(1, 20))
            
            table_data = [["ID", "Username", "Email", "Fecha Registro"]]
            for row in datos:
                table_data.append([str(row[0]), row[1], row[2], str(row[3])])
            
            table = Table(table_data, colWidths=[0.8*inch, 1.5*inch, 2.5*inch, 1.5*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#D4A5A5')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 9),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 30))
            
            footer = Paragraph(
                f"Total de usuarios: {len(datos)} | Diario de Emociones v1.0",
                fecha_style
            )
            elements.append(footer)
            
            doc.build(elements)
            messagebox.showinfo("Éxito", f"PDF generado correctamente")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar PDF: {e}")

    def exportar_entradas_pdf(self, fecha_inicio=None, fecha_fin=None, usuario_id=None):
        """Exporta entradas a PDF con filtros"""
        try:
            cursor = self.conn.cursor()
            
            query = """
                SELECT e.id, u.username, e.fecha, e.texto 
                FROM entradas e 
                JOIN usuarios u ON e.usuario_id = u.id 
                WHERE 1=1
            """
            params = []
            
            if fecha_inicio:
                query += " AND e.fecha >= %s"
                params.append(fecha_inicio)
            if fecha_fin:
                query += " AND e.fecha <= %s"
                params.append(fecha_fin)
            if usuario_id:
                query += " AND e.usuario_id = %s"
                params.append(usuario_id)
            
            query += " ORDER BY e.fecha DESC"
            
            cursor.execute(query, params)
            datos = cursor.fetchall()
            
            if not datos:
                messagebox.showinfo("Sin datos", "No hay entradas para exportar")
                return
            
            archivo = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=f"entradas_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            
            if not archivo:
                return
            
            doc = SimpleDocTemplate(archivo, pagesize=A4)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=22,
                textColor=colors.HexColor('#B86D6D'),
                spaceAfter=20,
                alignment=1
            )
            title = Paragraph("📖 Reporte de Entradas del Diario", title_style)
            elements.append(title)
            
            if fecha_inicio or fecha_fin or usuario_id:
                filtros_texto = "Filtros aplicados: "
                if fecha_inicio:
                    filtros_texto += f"Desde {fecha_inicio} "
                if fecha_fin:
                    filtros_texto += f"Hasta {fecha_fin} "
                if usuario_id:
                    filtros_texto += f"Usuario ID: {usuario_id}"
                
                filtro_para = Paragraph(filtros_texto, styles['Normal'])
                elements.append(filtro_para)
            
            elements.append(Spacer(1, 20))
            
            for entrada in datos:
                entrada_id, username, fecha, texto = entrada
                
                entrada_style = ParagraphStyle(
                    'Entrada',
                    parent=styles['Normal'],
                    fontSize=10,
                    leading=14,
                    spaceBefore=10,
                    spaceAfter=10
                )
                
                header_text = f"<b>Entrada #{entrada_id}</b> | Usuario: {username} | Fecha: {fecha}"
                elements.append(Paragraph(header_text, entrada_style))
                
                texto_truncado = texto[:500] + "..." if len(texto) > 500 else texto
                elements.append(Paragraph(texto_truncado, styles['Normal']))
                elements.append(Spacer(1, 15))
            
            doc.build(elements)
            messagebox.showinfo("Éxito", f"PDF generado con {len(datos)} entradas")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar PDF: {e}")

    def mostrar_dialogo_filtros_exportacion(self):
        """Diálogo para configurar filtros de exportación"""
        dialog = tk.Toplevel(self.root)
        dialog.title("Filtros de Exportación")
        dialog.geometry("400x300")
        dialog.configure(bg="#faf3e0")
        dialog.transient(self.root)
        dialog.grab_set()
        
        tk.Label(dialog, text="Configurar Filtros de Exportación", 
                font=("Arial", 14, "bold"), bg="#faf3e0", fg="#b86d6d").pack(pady=15)
        
        frame = tk.Frame(dialog, bg="#f8f0e3", padx=20, pady=20)
        frame.pack(padx=20, pady=10, fill="both", expand=True)
        
        tk.Label(frame, text="Fecha Inicio:", bg="#f8f0e3", font=("Arial", 10)).grid(row=0, column=0, sticky="w", pady=8)
        fecha_inicio = DateEntry(frame, width=25, date_pattern="yyyy-mm-dd")
        fecha_inicio.grid(row=0, column=1, sticky="w", pady=8)
        
        tk.Label(frame, text="Fecha Fin:", bg="#f8f0e3", font=("Arial", 10)).grid(row=1, column=0, sticky="w", pady=8)
        fecha_fin = DateEntry(frame, width=25, date_pattern="yyyy-mm-dd")
        fecha_fin.grid(row=1, column=1, sticky="w", pady=8)
        
        tk.Label(frame, text="ID Usuario:", bg="#f8f0e3", font=("Arial", 10)).grid(row=2, column=0, sticky="w", pady=8)
        usuario_entry = tk.Entry(frame, width=27)
        usuario_entry.grid(row=2, column=1, sticky="w", pady=8)
        
        tk.Label(frame, text="Formato:", bg="#f8f0e3", font=("Arial", 10)).grid(row=3, column=0, sticky="w", pady=8)
        formato_var = tk.StringVar(value="excel")
        ttk.Radiobutton(frame, text="Excel", variable=formato_var, value="excel").grid(row=3, column=1, sticky="w")
        ttk.Radiobutton(frame, text="PDF", variable=formato_var, value="pdf").grid(row=4, column=1, sticky="w")
        
        def aplicar_filtros():
            f_inicio = fecha_inicio.get() if fecha_inicio.get() else None
            f_fin = fecha_fin.get() if fecha_fin.get() else None
            usr_id = usuario_entry.get() if usuario_entry.get() else None
            
            if formato_var.get() == "excel":
                self.exportar_entradas_excel(f_inicio, f_fin, usr_id)
            else:
                self.exportar_entradas_pdf(f_inicio, f_fin, usr_id)
            
            dialog.destroy()
        
        btn_frame = tk.Frame(dialog, bg="#faf3e0")
        btn_frame.pack(pady=15)
        
        tk.Button(btn_frame, text="📊 Exportar", bg="#88c9a1", fg="white", 
                 width=15, command=aplicar_filtros).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="❌ Cancelar", bg="#e77f67", fg="white", 
                 width=15, command=dialog.destroy).pack(side=tk.LEFT, padx=5)

    # ----------------------------------------
    # MÓDULO 1: USUARIOS
    # ----------------------------------------
    def crear_formulario_usuarios(self):
        titulo = tk.Label(self.tab_usuarios, text="👤 Gestión de Usuarios", font=("Arial", 16, "bold"), fg="#b86d6d",
                          bg="#faf3e0")
        titulo.pack(pady=(20, 10))

        form_frame = tk.Frame(self.tab_usuarios, bg="#f8f0e3", padx=20, pady=20, relief="groove", bd=1)
        form_frame.pack(pady=10, padx=20, fill="x")

        tk.Label(form_frame, text="ID Usuario:", bg="#f8f0e3", font=("Arial", 12)).grid(row=0, column=0, sticky="w", pady=5)
        self.usuario_id_entry = tk.Entry(form_frame, width=30)
        self.usuario_id_entry.grid(row=0, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Username:", bg="#f8f0e3", font=("Arial", 12)).grid(row=1, column=0, sticky="w", pady=5)
        self.username_entry = tk.Entry(form_frame, width=30)
        self.username_entry.grid(row=1, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Email:", bg="#f8f0e3", font=("Arial", 12)).grid(row=2, column=0, sticky="w", pady=5)
        self.email_entry = tk.Entry(form_frame, width=30)
        self.email_entry.grid(row=2, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Contraseña:", bg="#f8f0e3", font=("Arial", 12)).grid(row=3, column=0, sticky="w", pady=5)
        self.password_entry = tk.Entry(form_frame, width=30, show="*")
        self.password_entry.grid(row=3, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Imagen de Perfil:", bg="#f8f0e3", font=("Arial", 12)).grid(row=4, column=0, sticky="w", pady=5)
        self.imagen_path = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.imagen_path, width=30).grid(row=4, column=1, sticky="w", pady=5)
        tk.Button(form_frame, text="🖼️ Seleccionar", command=self.seleccionar_imagen, bg="#a2b9bc", fg="white").grid(row=4, column=2, sticky="w", pady=5, padx=5)
        
        # Preview de imagen
        self.preview_label_usuario = tk.Label(form_frame, bg="#f8f0e3", text="Vista previa", relief="sunken", width=20, height=8)
        self.preview_label_usuario.grid(row=5, column=1, pady=10)

        btn_frame = tk.Frame(self.tab_usuarios, bg="#faf3e0")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="💾 Guardar", bg="#88c9a1", fg="white", width=12,
                  command=self.guardar_usuario).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✏️ Actualizar", bg="#a2b9bc", fg="white", width=12,
                  command=self.actualizar_usuario).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🗑️ Eliminar", bg="#e77f67", fg="white", width=12,
                  command=self.eliminar_usuario).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🧹 Limpiar", bg="#f5c091", fg="#5a4a42", width=12,
                  command=self.limpiar_usuario).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🔄 Cargar Datos", bg="#a2b9bc", fg="white", width=12,
                  command=self.cargar_usuarios_tabla).pack(side=tk.LEFT, padx=5)

        export_frame = tk.Frame(self.tab_usuarios, bg="#faf3e0")
        export_frame.pack(pady=5)
        
        tk.Button(export_frame, text="📊 Exportar Excel", bg="#4CAF50", fg="white", width=15,
                  command=self.exportar_usuarios_excel).pack(side=tk.LEFT, padx=5)
        tk.Button(export_frame, text="📄 Exportar PDF", bg="#2196F3", fg="white", width=15,
                  command=self.exportar_usuarios_pdf).pack(side=tk.LEFT, padx=5)

        tabla_frame = tk.Frame(self.tab_usuarios, bg="#faf3e0")
        tabla_frame.pack(pady=20, padx=20, fill="both", expand=True)

        tk.Label(tabla_frame, text="📋 Lista de Usuarios", font=("Arial", 14, "bold"),
                 fg="#b86d6d", bg="#faf3e0").pack(pady=(0, 10))

        columns = ("ID", "Username", "Email", "Fecha Registro")
        self.tabla_usuarios = ttk.Treeview(tabla_frame, columns=columns, show="headings", height=8)

        self.tabla_usuarios.heading("ID", text="ID")
        self.tabla_usuarios.heading("Username", text="Username")
        self.tabla_usuarios.heading("Email", text="Email")
        self.tabla_usuarios.heading("Fecha Registro", text="Fecha Registro")

        self.tabla_usuarios.column("ID", width=50)
        self.tabla_usuarios.column("Username", width=120)
        self.tabla_usuarios.column("Email", width=150)
        self.tabla_usuarios.column("Fecha Registro", width=120)

        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla_usuarios.yview)
        self.tabla_usuarios.configure(yscrollcommand=scrollbar.set)

        self.tabla_usuarios.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tabla_usuarios.bind("<<TreeviewSelect>>", self.seleccionar_usuario_tabla)
        self.cargar_usuarios_tabla()

    # FUNCIONES USUARIO (USANDO STORED PROCEDURES)
    def guardar_usuario(self):
        if not self.validar_id_usuario():
            return

        username = self.username_entry.get()
        email = self.email_entry.get()
        password = self.password_entry.get()
        imagen = self.imagen_path.get()

        if not self.validar_texto_username(username):
            return
        if not self.validar_email_usuario(email):
            return
        if not self.validar_texto_password(password):
            return
        if imagen and not self.validar_imagen(imagen):
            return

        if not messagebox.askyesno("Confirmar", f"¿Guardar usuario {username}?"):
            return

        try:
            # Procesar imagen si existe
            imagen_guardada = None
            if imagen:
                if self.validar_imagen(imagen):
                    # Copiar imagen a una carpeta de imágenes
                    import shutil
                    import os
                    img_dir = "imagenes/usuarios"
                    os.makedirs(img_dir, exist_ok=True)
                    extension = os.path.splitext(imagen)[1]
                    nuevo_nombre = f"{username}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
                    ruta_destino = os.path.join(img_dir, nuevo_nombre)
                    shutil.copy2(imagen, ruta_destino)
                    imagen_guardada = ruta_destino
                    print(f"✅ Imagen guardada en: {imagen_guardada}")
            
            cursor = self.conn.cursor()
            cursor.callproc('sp_crear_usuario', (username, email, f"hash_{password}", imagen_guardada))
            self.conn.commit()
            messagebox.showinfo("Éxito", f"Usuario {username} guardado exitosamente")
            self.limpiar_usuario()
            self.cargar_usuarios_tabla()
        except Error as e:
            messagebox.showerror("Error", f"Error al guardar usuario {username}: {e}")
            self.conn.rollback()

    def actualizar_usuario(self):
        user_id = self.usuario_id_entry.get()
        if not user_id:
            messagebox.showwarning("Advertencia", "ID de usuario es obligatorio para actualizar")
            return
        if not self.validar_id_usuario():
            return

        username = self.username_entry.get()
        email = self.email_entry.get()
        password = self.password_entry.get()

        if not self.validar_texto_username(username):
            return
        if not self.validar_email_usuario(email):
            return
        if password and not self.validar_texto_password(password):
            return

        if not messagebox.askyesno("Confirmar", f"¿Actualizar usuario con ID {user_id}?"):
            return

        try:
            cursor = self.conn.cursor()
            password_hash = f"hash_{password}" if password else "hash_sin_cambio"
            cursor.callproc('sp_actualizar_usuario', (user_id, username, email, password_hash))
            self.conn.commit()
            messagebox.showinfo("Éxito", f"Usuario con ID {user_id} actualizado exitosamente")
            self.limpiar_usuario()
            self.cargar_usuarios_tabla()
        except Error as e:
            messagebox.showerror("Error", f"Error al actualizar usuario {user_id}: {e}")
            self.conn.rollback()

    def eliminar_usuario(self):
        user_id = self.usuario_id_entry.get()
        if not user_id:
            messagebox.showwarning("Advertencia", "ID de usuario es obligatorio")
            return

        if messagebox.askyesno("Confirmar", "¿Estás seguro de eliminar a este usuario?"):
            try:
                cursor = self.conn.cursor()
                cursor.callproc('sp_eliminar_usuario', (user_id,))
                self.conn.commit()
                messagebox.showinfo("Éxito", f"Usuario {user_id} eliminado exitosamente")
                self.limpiar_usuario()
                self.cargar_usuarios_tabla()
            except Error as e:
                messagebox.showerror("Error", f"Error al eliminar usuario: {e}")
                self.conn.rollback()

    def limpiar_usuario(self):
        self.usuario_id_entry.delete(0, tk.END)
        self.username_entry.delete(0, tk.END)
        self.email_entry.delete(0, tk.END)
        self.password_entry.delete(0, tk.END)
        self.imagen_path.set("")
        self.preview_label_usuario.config(image='', text="Vista previa")

    # ----------------------------------------
    # MÓDULO 2: EMOCIONES
    # ----------------------------------------
    def crear_formulario_emociones(self):
        titulo = tk.Label(self.tab_emociones, text="😊 Catálogo de Emociones", font=("Arial", 16, "bold"), fg="#b86d6d",
                          bg="#faf3e0")
        titulo.pack(pady=(20, 10))

        form_frame = tk.Frame(self.tab_emociones, bg="#f8f0e3", padx=20, pady=20, relief="groove", bd=1)
        form_frame.pack(pady=10, padx=20, fill="x")

        tk.Label(form_frame, text="ID Emoción:", bg="#f8f0e3", font=("Arial", 12)).grid(row=0, column=0, sticky="w", pady=5)
        self.emocion_id_entry = tk.Entry(form_frame, width=30)
        self.emocion_id_entry.grid(row=0, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Nombre:", bg="#f8f0e3", font=("Arial", 12)).grid(row=1, column=0, sticky="w", pady=5)
        self.nombre_emocion_entry = tk.Entry(form_frame, width=30)
        self.nombre_emocion_entry.grid(row=1, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Emoji:", bg="#f8f0e3", font=("Arial", 12)).grid(row=2, column=0, sticky="w", pady=5)
        self.emoji_emocion_entry = tk.Entry(form_frame, width=30)
        self.emoji_emocion_entry.grid(row=2, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Imagen de Emoción:", bg="#f8f0e3", font=("Arial", 12)).grid(row=2, column=0, sticky="w", pady=5)
        self.imagen_emocion_path = tk.StringVar()
        tk.Entry(form_frame, textvariable=self.imagen_emocion_path, width=30).grid(row=2, column=1, sticky="w", pady=5)
        tk.Button(form_frame, text="🖼️ Seleccionar", command=self.seleccionar_imagen_emocion, bg="#88c9a1", fg="white").grid(row=2, column=2, sticky="w", padx=5, pady=5)
        
        # Preview de imagen de emoción
        self.preview_label_emocion = tk.Label(form_frame, bg="#f8f0e3", text="Vista previa", relief="sunken", width=20, height=8)
        self.preview_label_emocion.grid(row=3, column=1, pady=10)

        btn_frame = tk.Frame(self.tab_emociones, bg="#faf3e0")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="💾 Guardar", bg="#88c9a1", fg="white", width=12,
                  command=self.guardar_emocion).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✏️ Actualizar", bg="#a2b9bc", fg="white", width=12,
                  command=self.actualizar_emocion).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🗑️ Eliminar", bg="#e77f67", fg="white", width=12,
                  command=self.eliminar_emocion).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🧹 Limpiar", bg="#f5c091", fg="#5a4a42", width=12,
                  command=self.limpiar_emocion).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🔄 Cargar Datos", bg="#a2b9bc", fg="white", width=12,
                  command=self.cargar_emociones_tabla).pack(side=tk.LEFT, padx=5)

        export_frame = tk.Frame(self.tab_emociones, bg="#faf3e0")
        export_frame.pack(pady=5)
        
        tk.Button(export_frame, text="📊 Exportar Excel", bg="#4CAF50", fg="white", width=15,
                  command=self.exportar_emociones_excel).pack(side=tk.LEFT, padx=5)

        tabla_frame = tk.Frame(self.tab_emociones, bg="#faf3e0")
        tabla_frame.pack(pady=20, padx=20, fill="both", expand=True)

        tk.Label(tabla_frame, text="📋 Catálogo de Emociones", font=("Arial", 14, "bold"),
                 fg="#b86d6d", bg="#faf3e0").pack(pady=(0, 10))

        columns = ("ID", "Nombre", "Emoji")
        self.tabla_emociones = ttk.Treeview(tabla_frame, columns=columns, show="headings", height=8)

        self.tabla_emociones.heading("ID", text="ID")
        self.tabla_emociones.heading("Nombre", text="Nombre")
        self.tabla_emociones.heading("Emoji", text="Emoji")

        self.tabla_emociones.column("ID", width=50)
        self.tabla_emociones.column("Nombre", width=150)
        self.tabla_emociones.column("Emoji", width=80)

        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla_emociones.yview)
        self.tabla_emociones.configure(yscrollcommand=scrollbar.set)

        self.tabla_emociones.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tabla_emociones.bind("<<TreeviewSelect>>", self.seleccionar_emocion_tabla)
        self.cargar_emociones_tabla()

    # FUNCIONES EMOCIONES (USANDO STORED PROCEDURES)
    def guardar_emocion(self):
        if not self.validar_id_emocion():
            return

        nombre = self.nombre_emocion_entry.get()
        emoji = self.emoji_emocion_entry.get()
        imagen = self.imagen_emocion_path.get().strip()
        
        if imagen and not self.validar_imagen(imagen):
            return

        if not self.validar_texto_nombre_emocion(nombre):
            return

        if not messagebox.askyesno("Confirmar", f"¿Guardar emoción {nombre}?"):
            return

        try:
            # Procesar y guardar imagen si existe
            imagen_guardada = None
            if imagen:
                if self.validar_imagen(imagen):
                    # Copiar imagen a una carpeta de imágenes
                    import os
                    import shutil
                    img_dir = "imagenes/emociones"
                    os.makedirs(img_dir, exist_ok=True)
                    extension = os.path.splitext(imagen)[1]
                    nuevo_nombre = f"emocion_{nombre}_{datetime.now().strftime('%Y%m%d_%H%M%S')}{extension}"
                    ruta_destino = os.path.join(img_dir, nuevo_nombre)
                    shutil.copy2(imagen, ruta_destino)
                    imagen_guardada = ruta_destino
                    print(f"✅ Imagen de emoción guardada en: {imagen_guardada}")
            
            cursor = self.conn.cursor()
            cursor.callproc('sp_crear_emocion', (nombre, emoji, imagen_guardada))
            self.conn.commit()
            messagebox.showinfo("Éxito", f"Emoción {nombre} guardada exitosamente")
            self.limpiar_emocion()
            self.cargar_emociones_tabla()
        except Error as e:
            messagebox.showerror("Error", f"Error al guardar emoción: {e}")
            self.conn.rollback()

    def actualizar_emocion(self):
        emocion_id = self.emocion_id_entry.get()
        if not emocion_id:
            messagebox.showwarning("Advertencia", "ID de la emoción es obligatorio")
            return
        if not self.validar_id_emocion():
            return

        nombre = self.nombre_emocion_entry.get()

        if not self.validar_texto_nombre_emocion(nombre):
            return

        if not messagebox.askyesno("Confirmar", f"¿Actualizar emoción con ID {emocion_id}?"):
            return

        try:
            cursor = self.conn.cursor()
            cursor.callproc('sp_actualizar_emocion', (emocion_id, nombre))
            self.conn.commit()
            messagebox.showinfo("Éxito", f"Emoción {emocion_id} actualizada exitosamente")
            self.limpiar_emocion()
            self.cargar_emociones_tabla()
        except Error as e:
            messagebox.showerror("Error", f"Error al actualizar emoción: {e}")
            self.conn.rollback()

    def eliminar_emocion(self):
        emocion_id = self.emocion_id_entry.get()
        if not emocion_id:
            messagebox.showwarning("Advertencia", "ID de la emoción es obligatorio")
            return
        if not self.validar_id_emocion():
            return

        if messagebox.askyesno("Confirmar", "¿Estás seguro de eliminar esta emoción?"):
            try:
                cursor = self.conn.cursor()
                cursor.callproc('sp_eliminar_emocion', (emocion_id,))
                self.conn.commit()
                messagebox.showinfo("Éxito", f"Emoción con ID {emocion_id} eliminada exitosamente")
                self.limpiar_emocion()
                self.cargar_emociones_tabla()
            except Error as e:
                messagebox.showerror("Error", f"Error al eliminar emoción: {e}")
                self.conn.rollback()

    def limpiar_emocion(self):
        self.emocion_id_entry.delete(0, tk.END)
        self.nombre_emocion_entry.delete(0, tk.END)
        self.emoji_emocion_entry.delete(0, tk.END)
        self.imagen_emocion_path.set("")
        self.preview_label_emocion.config(image='', text="Vista previa")

    # ----------------------------------------
    # MÓDULO 3: ENTRADAS
    # ----------------------------------------
    def crear_formulario_entradas(self):
        titulo = tk.Label(self.tab_entradas, text="📖 Entradas del Diario", font=("Arial", 16, "bold"), fg="#b86d6d",
                          bg="#faf3e0")
        titulo.pack(pady=(20, 10))

        form_frame = tk.Frame(self.tab_entradas, bg="#f8f0e3", padx=20, pady=20, relief="groove", bd=1)
        form_frame.pack(pady=10, padx=20, fill="x")

        tk.Label(form_frame, text="ID Entrada:", bg="#f8f0e3", font=("Arial", 12)).grid(row=0, column=0, sticky="w", pady=5)
        self.entrada_id_entry = tk.Entry(form_frame, width=30)
        self.entrada_id_entry.grid(row=0, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Usuario ID:", bg="#f8f0e3", font=("Arial", 12)).grid(row=1, column=0, sticky="w", pady=5)
        self.entrada_usuario_id_entry = tk.Entry(form_frame, width=30)
        self.entrada_usuario_id_entry.grid(row=1, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Fecha:", bg="#f8f0e3", font=("Arial", 12)).grid(row=2, column=0, sticky="w", pady=5)
        self.fecha_entry = DateEntry(form_frame, width=28, date_pattern="yyyy-mm-dd")
        self.fecha_entry.grid(row=2, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Texto:", bg="#f8f0e3", font=("Arial", 12)).grid(row=3, column=0, sticky="nw", pady=5)
        self.texto_entry = tk.Text(form_frame, width=30, height=5)
        self.texto_entry.grid(row=3, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Emociones (IDs):", bg="#f8f0e3", font=("Arial", 12)).grid(row=4, column=0, sticky="w", pady=5)
        self.emociones_ids_entry = tk.Entry(form_frame, width=30)
        self.emociones_ids_entry.grid(row=4, column=1, sticky="w", pady=5)

        btn_frame = tk.Frame(self.tab_entradas, bg="#faf3e0")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="💾 Guardar", bg="#88c9a1", fg="white", width=12,
                  command=self.guardar_entrada).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="✏️ Actualizar", bg="#a2b9bc", fg="white", width=12,
                  command=self.actualizar_entrada).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🗑️ Eliminar", bg="#e77f67", fg="white", width=12,
                  command=self.eliminar_entrada).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🧹 Limpiar", bg="#f5c091", fg="#5a4a42", width=12,
                  command=self.limpiar_entrada).pack(side=tk.LEFT, padx=5)
        tk.Button(btn_frame, text="🔄 Cargar Datos", bg="#a2b9bc", fg="white", width=12,
                  command=self.cargar_entradas_tabla).pack(side=tk.LEFT, padx=5)

        export_frame = tk.Frame(self.tab_entradas, bg="#faf3e0")
        export_frame.pack(pady=5)
        
        tk.Button(export_frame, text="📊 Exportar con Filtros", bg="#FF9800", fg="white", width=20,
                  command=self.mostrar_dialogo_filtros_exportacion).pack(side=tk.LEFT, padx=5)

        tabla_frame = tk.Frame(self.tab_entradas, bg="#faf3e0")
        tabla_frame.pack(pady=20, padx=20, fill="both", expand=True)

        tk.Label(tabla_frame, text="📋 Historial de Entradas", font=("Arial", 14, "bold"),
                 fg="#b86d6d", bg="#faf3e0").pack(pady=(0, 10))

        columns = ("ID", "Usuario", "Fecha", "Resumen")
        self.tabla_entradas = ttk.Treeview(tabla_frame, columns=columns, show="headings", height=8)

        self.tabla_entradas.heading("ID", text="ID")
        self.tabla_entradas.heading("Usuario", text="Usuario")
        self.tabla_entradas.heading("Fecha", text="Fecha")
        self.tabla_entradas.heading("Resumen", text="Resumen")

        self.tabla_entradas.column("ID", width=50)
        self.tabla_entradas.column("Usuario", width=100)
        self.tabla_entradas.column("Fecha", width=100)
        self.tabla_entradas.column("Resumen", width=200)

        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla_entradas.yview)
        self.tabla_entradas.configure(yscrollcommand=scrollbar.set)

        self.tabla_entradas.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        self.tabla_entradas.bind("<<TreeviewSelect>>", self.seleccionar_entrada_tabla)
        self.cargar_entradas_tabla()

    # FUNCIONES ENTRADAS (USANDO STORED PROCEDURES)
    def guardar_entrada(self):
        if not self.validar_usuario_id_entrada():
            return

        usuario_id = self.entrada_usuario_id_entry.get()
        texto = self.texto_entry.get("1.0", tk.END).strip()
        emociones_ids = self.emociones_ids_entry.get()

        if not self.validar_entrada(texto):
            return

        if not messagebox.askyesno("Confirmar", f"¿Guardar entrada para Usuario {usuario_id}?"):
            return

        try:
            cursor = self.conn.cursor()
            fecha = self.fecha_entry.get()

            cursor.callproc('sp_crear_entrada', (usuario_id, fecha, texto))
            
            # Obtener el ID de la entrada recién creada
            cursor.execute("SELECT LAST_INSERT_ID()")
            entrada_id = cursor.fetchone()[0]

            if emociones_ids:
                emociones_lista = [eid.strip() for eid in emociones_ids.split(',')]
                for emocion_id in emociones_lista:
                    if emocion_id and emocion_id.isdigit():
                        query_relacion = "INSERT INTO entrada_emocion (entrada_id, emocion_id) VALUES (%s, %s)"
                        cursor.execute(query_relacion, (entrada_id, emocion_id))

            self.conn.commit()
            messagebox.showinfo("Éxito", f"Entrada guardada para Usuario ID: {usuario_id}")
            self.limpiar_entrada()
            self.cargar_entradas_tabla()
        except Error as e:
            messagebox.showerror("Error", f"No se pudo guardar: {e}")
            self.conn.rollback()

    def actualizar_entrada(self):
        entrada_id = self.entrada_id_entry.get()
        if not entrada_id:
            messagebox.showwarning("Advertencia", "ID de entrada es obligatorio")
            return
        if not self.validar_id_entrada():
            return

        usuario_id = self.entrada_usuario_id_entry.get()
        texto = self.texto_entry.get("1.0", tk.END).strip()

        if not self.validar_usuario_id_entrada():
            return
        if not self.validar_entrada(texto):
            return

        if not messagebox.askyesno("Confirmar", f"¿Actualizar entrada con ID {entrada_id}?"):
            return

        try:
            cursor = self.conn.cursor()
            fecha = self.fecha_entry.get()
            cursor.callproc('sp_actualizar_entrada', (entrada_id, usuario_id, fecha, texto))
            self.conn.commit()
            messagebox.showinfo("Éxito", f"Entrada con ID: {entrada_id} actualizada correctamente")
            self.limpiar_entrada()
            self.cargar_entradas_tabla()
        except Error as e:
            messagebox.showerror("Error", f"No se pudo actualizar: {e}")
            self.conn.rollback()

    def eliminar_entrada(self):
        entrada_id = self.entrada_id_entry.get()
        if not entrada_id:
            messagebox.showwarning("Advertencia", "ID de entrada es obligatorio")
            return
        if not self.validar_id_entrada():
            return

        if messagebox.askyesno("Confirmar", "¿Estás seguro de eliminar esta entrada?"):
            try:
                cursor = self.conn.cursor()
                cursor.callproc('sp_eliminar_entrada', (entrada_id,))
                self.conn.commit()
                messagebox.showinfo("Éxito", f"Entrada con ID {entrada_id} eliminada correctamente")
                self.limpiar_entrada()
                self.cargar_entradas_tabla()
            except Error as e:
                messagebox.showerror("Error", f"No se pudo eliminar: {e}")
                self.conn.rollback()

    def limpiar_entrada(self):
        self.entrada_id_entry.delete(0, tk.END)
        self.entrada_usuario_id_entry.delete(0, tk.END)
        self.fecha_entry.delete(0, tk.END)
        self.texto_entry.delete("1.0", tk.END)
        self.emociones_ids_entry.delete(0, tk.END)

    # ----------------------------------------
    # MÓDULO 4: REPORTES
    # ----------------------------------------
    def crear_formulario_reportes(self):
        titulo = tk.Label(self.tab_reportes, text="📈 Resumen Emocional", font=("Arial", 16, "bold"), fg="#b86d6d",
                          bg="#faf3e0")
        titulo.pack(pady=(20, 10))

        form_frame = tk.Frame(self.tab_reportes, bg="#f8f0e3", padx=20, pady=20, relief="groove", bd=1)
        form_frame.pack(pady=10, padx=20, fill="x")

        tk.Label(form_frame, text="Selecciona un usuario (ID):", bg="#f8f0e3", font=("Arial", 12)).grid(row=0, column=0,
                                                                                                        sticky="w",
                                                                                                        pady=5)
        self.reporte_usuario_id_entry = tk.Entry(form_frame, width=30)
        self.reporte_usuario_id_entry.grid(row=0, column=1, sticky="w", pady=5)

        tk.Label(form_frame, text="Período:", bg="#f8f0e3", font=("Arial", 12)).grid(row=1, column=0, sticky="w",
                                                                                     pady=5)
        self.periodo_var = tk.StringVar(value="semana")
        ttk.Radiobutton(form_frame, text="Semana", variable=self.periodo_var, value="semana").grid(row=1, column=1,
                                                                                                   sticky="w")
        ttk.Radiobutton(form_frame, text="Mes", variable=self.periodo_var, value="mes").grid(row=1, column=2,
                                                                                             sticky="w")

        btn_frame = tk.Frame(self.tab_reportes, bg="#faf3e0")
        btn_frame.pack(pady=20)

        tk.Button(btn_frame, text="📊 Generar Reporte", bg="#88c9a1", fg="white", width=20,
                  command=self.generar_reporte).pack(padx=5)

        export_frame = tk.Frame(self.tab_reportes, bg="#faf3e0")
        export_frame.pack(pady=5)
        
        tk.Button(export_frame, text="📊 Exportar Reporte Excel", bg="#4CAF50", fg="white", width=20,
                  command=self.exportar_reporte_excel).pack(side=tk.LEFT, padx=5)
        tk.Button(export_frame, text="📄 Exportar Reporte PDF", bg="#2196F3", fg="white", width=20,
                  command=self.exportar_reporte_pdf).pack(side=tk.LEFT, padx=5)

        tabla_frame = tk.Frame(self.tab_reportes, bg="#faf3e0")
        tabla_frame.pack(pady=20, padx=20, fill="both", expand=True)

        tk.Label(tabla_frame, text="📊 Estadísticas Emocionales", font=("Arial", 14, "bold"),
                 fg="#b86d6d", bg="#faf3e0").pack(pady=(0, 10))

        columns = ("Emoción", "Frecuencia", "Última Registro")
        self.tabla_reportes = ttk.Treeview(tabla_frame, columns=columns, show="headings", height=8)

        self.tabla_reportes.heading("Emoción", text="Emoción")
        self.tabla_reportes.heading("Frecuencia", text="Frecuencia")
        self.tabla_reportes.heading("Última Registro", text="Última Registro")

        self.tabla_reportes.column("Emoción", width=150)
        self.tabla_reportes.column("Frecuencia", width=100)
        self.tabla_reportes.column("Última Registro", width=120)

        scrollbar = ttk.Scrollbar(tabla_frame, orient="vertical", command=self.tabla_reportes.yview)
        self.tabla_reportes.configure(yscrollcommand=scrollbar.set)

        self.tabla_reportes.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

    def generar_reporte(self):
        usuario_id = self.reporte_usuario_id_entry.get()
        periodo = self.periodo_var.get()

        if not usuario_id:
            messagebox.showwarning("Advertencia", "Por favor, Ingrese un ID de usuario para generar reporte")
            return
        if not usuario_id.isdigit():
            messagebox.showerror("Error", "ID de usuario debe ser un número")
            return

        try:
            cursor = self.conn.cursor()

            query = """
            SELECT e.nombre, COUNT(ee.emocion_id) as frecuencia, MAX(en.fecha) as ultima_fecha
            FROM entrada_emocion ee 
            JOIN emociones e ON ee.emocion_id = e.id 
            JOIN entradas en ON ee.entrada_id = en.id 
            WHERE en.usuario_id = %s 
            GROUP BY e.nombre 
            ORDER BY frecuencia DESC
            """

            cursor.execute(query, (usuario_id,))
            resultados = cursor.fetchall()

            for item in self.tabla_reportes.get_children():
                self.tabla_reportes.delete(item)

            if resultados:
                for emocion, frecuencia, ultima_fecha in resultados:
                    self.tabla_reportes.insert("", "end", values=(emocion, f"{frecuencia} veces", ultima_fecha))
                messagebox.showinfo("Éxito", f"Reporte generado para usuario {usuario_id}")
            else:
                messagebox.showinfo("Reporte", "No hay datos para generar el reporte")

        except Error as e:
            messagebox.showerror("Error", f"Error al generar reporte: {e}")

    def exportar_reporte_excel(self):
        """Exporta el reporte actual a Excel"""
        usuario_id = self.reporte_usuario_id_entry.get()
        
        if not usuario_id:
            messagebox.showwarning("Advertencia", "Genera primero un reporte")
            return
        
        items = self.tabla_reportes.get_children()
        if not items:
            messagebox.showinfo("Sin datos", "No hay datos en el reporte para exportar")
            return
        
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".xlsx",
                filetypes=[("Excel files", "*.xlsx")],
                initialfile=f"reporte_emocional_usuario_{usuario_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            )
            
            if not archivo:
                return
            
            wb = Workbook()
            ws = wb.active
            ws.title = f"Reporte Usuario {usuario_id}"
            
            ws.merge_cells('A1:C1')
            titulo_cell = ws['A1']
            titulo_cell.value = f"Reporte Emocional - Usuario ID: {usuario_id}"
            titulo_cell.font = Font(bold=True, size=16, color="B86D6D")
            titulo_cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.merge_cells('A2:C2')
            fecha_cell = ws['A2']
            fecha_cell.value = f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}"
            fecha_cell.font = Font(italic=True, size=10)
            fecha_cell.alignment = Alignment(horizontal="center")
            
            header_fill = PatternFill(start_color="E77F67", end_color="E77F67", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            
            headers = ["Emoción", "Frecuencia", "Último Registro"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=4, column=col, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            for row_idx, item in enumerate(items, 5):
                values = self.tabla_reportes.item(item)['values']
                for col_idx, value in enumerate(values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.alignment = Alignment(horizontal="left", vertical="center")
            
            ws.column_dimensions['A'].width = 25
            ws.column_dimensions['B'].width = 20
            ws.column_dimensions['C'].width = 20
            
            total_registros = len(items)
            ws.cell(row=len(items)+6, column=1, value="Total de emociones registradas:").font = Font(bold=True)
            ws.cell(row=len(items)+6, column=2, value=total_registros)
            
            wb.save(archivo)
            messagebox.showinfo("Éxito", f"Reporte exportado correctamente")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar reporte: {e}")

    def exportar_reporte_pdf(self):
        """Exporta el reporte actual a PDF"""
        usuario_id = self.reporte_usuario_id_entry.get()
        
        if not usuario_id:
            messagebox.showwarning("Advertencia", "Genera primero un reporte")
            return
        
        items = self.tabla_reportes.get_children()
        if not items:
            messagebox.showinfo("Sin datos", "No hay datos en el reporte para exportar")
            return
        
        try:
            archivo = filedialog.asksaveasfilename(
                defaultextension=".pdf",
                filetypes=[("PDF files", "*.pdf")],
                initialfile=f"reporte_emocional_usuario_{usuario_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            )
            
            if not archivo:
                return
            
            doc = SimpleDocTemplate(archivo, pagesize=letter)
            elements = []
            styles = getSampleStyleSheet()
            
            title_style = ParagraphStyle(
                'CustomTitle',
                parent=styles['Heading1'],
                fontSize=24,
                textColor=colors.HexColor('#B86D6D'),
                spaceAfter=20,
                alignment=1
            )
            title = Paragraph(f"📊 Reporte Emocional - Usuario ID: {usuario_id}", title_style)
            elements.append(title)
            
            fecha_style = ParagraphStyle(
                'Fecha',
                parent=styles['Normal'],
                fontSize=11,
                textColor=colors.grey,
                alignment=1,
                spaceAfter=30
            )
            fecha = Paragraph(f"Generado el: {datetime.now().strftime('%d/%m/%Y %H:%M:%S')}<br/>Período: {self.periodo_var.get().capitalize()}", fecha_style)
            elements.append(fecha)
            
            table_data = [["Emoción", "Frecuencia", "Último Registro"]]
            for item in items:
                values = self.tabla_reportes.item(item)['values']
                table_data.append([str(values[0]), str(values[1]), str(values[2])])
            
            table = Table(table_data, colWidths=[2.5*inch, 2*inch, 2*inch])
            table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E77F67')),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                ('FONTSIZE', (0, 1), (-1, -1), 10),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ]))
            
            elements.append(table)
            elements.append(Spacer(1, 30))
            
            resumen_style = ParagraphStyle(
                'Resumen',
                parent=styles['Normal'],
                fontSize=12,
                textColor=colors.HexColor('#5A4A42'),
                spaceAfter=10
            )
            
            total_emociones = len(items)
            elements.append(Paragraph(f"<b>Resumen Estadístico:</b>", resumen_style))
            elements.append(Paragraph(f"• Total de emociones diferentes registradas: {total_emociones}", resumen_style))
            
            if items:
                primera_emocion = self.tabla_reportes.item(items[0])['values']
                elements.append(Paragraph(f"• Emoción más frecuente: {primera_emocion[0]} ({primera_emocion[1]})", resumen_style))
            
            elements.append(Spacer(1, 30))
            
            footer = Paragraph(
                "Diario de Emociones v1.0 - Tu espacio seguro para sentir 🌿",
                fecha_style
            )
            elements.append(footer)
            
            doc.build(elements)
            messagebox.showinfo("Éxito", f"Reporte PDF generado correctamente")
            
        except Exception as e:
            messagebox.showerror("Error", f"Error al exportar PDF: {e}")

    def crear_footer(self):
        footer = tk.Label(self.root, text="Diario de Emociones v1.0 — Tu espacio seguro para sentir 🌿",
                          font=("Arial", 10, "italic"), fg="#7d6e65", bg="#faf3e0")
        footer.pack(side=tk.BOTTOM, pady=15)


# ----------------------------------------
# INICIAR LA APLICACIÓN
# ----------------------------------------
if __name__ == "__main__":
    root = tk.Tk()
    app = DiarioEmocionesApp(root)
    root.mainloop()
